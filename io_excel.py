"""
io_excel.py
-----------
Read network data from the PE80 Excel template and write solver results back.

Expected sheet names:  NODES, PIPES
Output sheet name:     RESULTS

All unit conversions happen here — the rest of the solver works in pure SI.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              PatternFill)
from openpyxl.utils import get_column_letter
from typing import List, Tuple, Optional
import logging

from network import Node, Pipe
from gas_properties import P_STD, T_STD

log = logging.getLogger(__name__)

# ── Unit conversion constants ─────────────────────────────────────────────────
KPA_TO_PA    = 1_000.0
SM3H_TO_SM3S = 1.0 / 3600.0
MM_TO_M      = 1.0 / 1_000.0


# ── Internal helpers ──────────────────────────────────────────────────────────

def _check_sheets(filepath: str, required: list) -> None:
    """Raise a clear ValueError if any required sheet is missing."""
    wb = load_workbook(filepath, read_only=True)
    found = wb.sheetnames
    wb.close()
    missing = [s for s in required if s not in found]
    if missing:
        raise ValueError(
            f"Required sheet(s) not found in '{filepath}': {missing}\n"
            f"Sheets present: {found}"
        )


def _find_header_row(filepath: str, sheet: str, id_col: str) -> int:
    """Scan rows to find the one containing id_col — handles any number of title rows."""
    df_raw = pd.read_excel(filepath, sheet_name=sheet, header=None, dtype=str)
    for i, row in df_raw.iterrows():
        if any(str(v).strip() == id_col for v in row.values):
            return int(i)
    raise ValueError(
        f"Cannot find a row with column '{id_col}' in sheet '{sheet}'. "
        f"Check that the column header is spelled exactly as '{id_col}'."
    )


def _require_cols(df: pd.DataFrame, required: list, sheet: str) -> None:
    """Raise a clear ValueError listing every missing required column."""
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"Sheet '{sheet}': missing required column(s): {missing}\n"
            f"Columns found: {list(df.columns)}"
        )


def _safe_float(val, default, label: str = '') -> float:
    """Convert val to float; return default (and warn) on non-numeric input."""
    if pd.isna(val) if not isinstance(val, str) else (val.strip() == '' or val.strip().lower() == 'nan'):
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        if label:
            log.warning(f"Non-numeric value in '{label}': {val!r} — using {default}")
        return default


def read_nodes(filepath: str,
               sheet: str = 'NODES',
               p_init_unknown: Optional[float] = None) -> List[Node]:
    """
    Read the NODES sheet from the PE80 Excel template.

    Required columns: Node_ID, Node_Type
    Optional columns: Name, Elevation_m, Known_P_kPa_g, Demand_Sm3h,
                      Diversity_Factor (or Div_Factor), Eff_Demand_Sm3h,
                      Min_P_kPa_g, X_Coord, Y_Coord
    """
    _check_sheets(filepath, [sheet])

    hdr = _find_header_row(filepath, sheet, 'Node_ID')
    df  = pd.read_excel(filepath, sheet_name=sheet, header=hdr)
    df.columns = df.columns.str.strip()

    _require_cols(df, ['Node_ID', 'Node_Type'], sheet)

    # Drop blank rows
    df = df[df['Node_ID'].astype(str).str.strip().str.len() > 0]
    df = df.dropna(subset=['Node_ID'])
    df = df[~df['Node_ID'].astype(str).str.strip().str.lower().isin(['nan', 'none', ''])]

    # Duplicate Node_ID check
    ids = df['Node_ID'].astype(str).str.strip()
    dupes = ids[ids.duplicated()].unique().tolist()
    if dupes:
        raise ValueError(f"Duplicate Node_ID(s) in sheet '{sheet}': {dupes}")

    # Accept both 'Diversity_Factor' and 'Div_Factor' column names
    if 'Diversity_Factor' not in df.columns and 'Div_Factor' in df.columns:
        df = df.rename(columns={'Div_Factor': 'Diversity_Factor'})

    VALID_TYPES = {'source', 'junction', 'offtake'}

    nodes = []
    for _, row in df.iterrows():
        nid   = str(row['Node_ID']).strip()
        ntype = str(row['Node_Type']).strip().lower()
        if ntype not in VALID_TYPES:
            log.warning(f"Node {nid}: unrecognised type '{ntype}' — treated as junction")
            ntype = 'junction'

        name_val  = row.get('Name', None)
        label     = str(name_val).strip() if pd.notna(name_val) else nid
        elevation = _safe_float(row.get('Elevation_m'), 0.0, f"Node {nid} Elevation_m")
        x_coord   = _safe_float(row.get('X_Coord'),    0.0, f"Node {nid} X_Coord")
        y_coord   = _safe_float(row.get('Y_Coord'),    0.0, f"Node {nid} Y_Coord")

        # Pressure — sources only
        if ntype == 'source':
            raw_p = row.get('Known_P_kPa_g', None)
            p_val = _safe_float(raw_p, None, f"Node {nid} Known_P_kPa_g")
            if p_val is None or p_val <= 0:
                raise ValueError(
                    f"Node {nid} (source): Known_P_kPa_g must be a positive number, "
                    f"got {raw_p!r}. Check the cell contains a plain number, not a formula."
                )
            pressure = p_val * KPA_TO_PA + P_STD
        else:
            pressure = 0.0

        # Demand — prefer Eff_Demand, fall back to Demand × DF
        if ntype == 'offtake':
            eff_dem = _safe_float(row.get('Eff_Demand_Sm3h'), None, f"Node {nid} Eff_Demand_Sm3h")
            if eff_dem is not None:
                demand = eff_dem * SM3H_TO_SM3S
            else:
                raw_dem = _safe_float(row.get('Demand_Sm3h'), None, f"Node {nid} Demand_Sm3h")
                if 'Diversity_Factor' in df.columns:
                    div_fac = _safe_float(row.get('Diversity_Factor'), 1.0, f"Node {nid} Diversity_Factor")
                else:
                    log.warning(f"Node {nid}: Diversity_Factor column missing — factor set to 1.0")
                    div_fac = 1.0
                demand = (raw_dem * div_fac * SM3H_TO_SM3S) if raw_dem is not None else 0.0
        else:
            demand = 0.0

        # Minimum pressure
        min_p_val = _safe_float(row.get('Min_P_kPa_g'), None, f"Node {nid} Min_P_kPa_g")
        min_pressure = (min_p_val * KPA_TO_PA + P_STD) if (min_p_val is not None and min_p_val > 0) else P_STD

        nodes.append(Node(
            id=nid, label=label, node_type=ntype,
            elevation=elevation, pressure=pressure, demand=demand,
            min_pressure=min_pressure,
            x_coord=x_coord, y_coord=y_coord,
        ))

    source_pressures = [n.pressure for n in nodes if n.is_source()]
    if not source_pressures:
        raise ValueError(
            f"No source nodes found in sheet '{sheet}'. "
            "Add at least one row with Node_Type='source' and a Known_P_kPa_g value."
        )

    p_init = p_init_unknown if p_init_unknown else min(source_pressures)
    for n in nodes:
        if not n.is_source():
            n.pressure = p_init

    log.info(f"Read {len(nodes)} nodes from '{sheet}' "
             f"({sum(1 for n in nodes if n.is_source())} sources, "
             f"{sum(1 for n in nodes if n.is_offtake())} offtakes, "
             f"{sum(1 for n in nodes if n.is_junction())} junctions)")
    return nodes


def read_pipes(filepath: str,
               sheet: str = 'PIPES',
               status_filter: str = 'open') -> List[Pipe]:
    """
    Read the PIPES sheet from the PE80 Excel template.

    Required columns: Pipe_ID, From_Node, To_Node
    Length   : Length_m  →  Length_m_SI  (first non-null wins)
    Diameter : ID_m_SI   →  ID_mm        →  OD_mm + SDR  (first available wins)
    Roughness: Rough_m_SI → Roughness_mm → PE80 default 0.007 mm
    """
    _check_sheets(filepath, [sheet])

    hdr = _find_header_row(filepath, sheet, 'Pipe_ID')
    df  = pd.read_excel(filepath, sheet_name=sheet, header=hdr)
    df.columns = df.columns.str.strip()

    _require_cols(df, ['Pipe_ID', 'From_Node', 'To_Node'], sheet)

    # Drop blank rows
    df = df[df['Pipe_ID'].astype(str).str.strip().str.len() > 0]
    df = df.dropna(subset=['Pipe_ID'])
    df = df[~df['Pipe_ID'].astype(str).str.strip().str.lower().isin(['nan', 'none', ''])]

    # Duplicate Pipe_ID check
    pids = df['Pipe_ID'].astype(str).str.strip()
    dupes = pids[pids.duplicated()].unique().tolist()
    if dupes:
        raise ValueError(f"Duplicate Pipe_ID(s) in sheet '{sheet}': {dupes}")

    # Normalise Status: strip whitespace + lowercase before filtering
    if 'Status' in df.columns:
        df['Status'] = df['Status'].astype(str).str.strip().str.lower()
        df = df[df['Status'] == status_filter.lower()]

    pipes = []
    for _, row in df.iterrows():
        pid       = str(row['Pipe_ID']).strip()
        from_node = str(row['From_Node']).strip()
        to_node   = str(row['To_Node']).strip()

        # ── Length ───────────────────────────────────────────────────────────
        length = None
        for col in ('Length_m', 'Length_m_SI'):
            if col in df.columns:
                v = _safe_float(row.get(col), None, f"Pipe {pid} {col}")
                if v is not None:
                    if col != 'Length_m':
                        log.debug(f"Pipe {pid}: length from fallback column '{col}'")
                    length = v
                    break
        if length is None:
            raise ValueError(f"Pipe {pid}: no valid length found (checked Length_m, Length_m_SI).")
        if length <= 0:
            raise ValueError(f"Pipe {pid}: length must be > 0, got {length} m.")

        # ── Diameter ─────────────────────────────────────────────────────────
        diameter = None
        if 'ID_m_SI' in df.columns:
            diameter = _safe_float(row.get('ID_m_SI'), None, f"Pipe {pid} ID_m_SI")
        if diameter is None and 'ID_mm' in df.columns:
            v = _safe_float(row.get('ID_mm'), None, f"Pipe {pid} ID_mm")
            if v is not None:
                diameter = v * MM_TO_M
        if diameter is None and 'OD_mm' in df.columns and 'SDR' in df.columns:
            od  = _safe_float(row.get('OD_mm'), None, f"Pipe {pid} OD_mm")
            sdr = _safe_float(row.get('SDR'),   None, f"Pipe {pid} SDR")
            if od is not None and sdr is not None:
                if sdr <= 2.0:
                    raise ValueError(
                        f"Pipe {pid}: SDR={sdr} is invalid (must be > 2). "
                        "Check the SDR cell — it may contain a formula or wrong value."
                    )
                diameter = od * (1.0 - 2.0 / sdr) * MM_TO_M
                log.debug(f"Pipe {pid}: diameter from OD/SDR = {od}mm / SDR{sdr} → {diameter*1000:.3f}mm ID")
        if diameter is None:
            raise ValueError(
                f"Pipe {pid}: no valid diameter found. "
                "Provide one of: ID_m_SI, ID_mm, or OD_mm + SDR columns."
            )
        if diameter <= 0:
            raise ValueError(f"Pipe {pid}: computed diameter ≤ 0 ({diameter:.6f} m) — check OD_mm and SDR.")

        # ── Roughness ────────────────────────────────────────────────────────
        roughness = None
        if 'Rough_m_SI' in df.columns:
            roughness = _safe_float(row.get('Rough_m_SI'), None, f"Pipe {pid} Rough_m_SI")
        if roughness is None and 'Roughness_mm' in df.columns:
            v = _safe_float(row.get('Roughness_mm'), None, f"Pipe {pid} Roughness_mm")
            if v is not None:
                roughness = v * MM_TO_M
        if roughness is None:
            roughness = 7e-6  # PE80 default
            log.debug(f"Pipe {pid}: roughness not found — using PE80 default (0.007 mm)")
        if roughness < 0:
            raise ValueError(f"Pipe {pid}: roughness must be ≥ 0, got {roughness} m.")

        status = str(row.get('Status', 'open')).strip().lower()

        pipes.append(Pipe(
            id=pid, from_node=from_node, to_node=to_node,
            length=length, diameter=diameter,
            roughness=roughness, status=status,
        ))

    log.info(f"Read {len(pipes)} '{status_filter}' pipes from '{sheet}'")
    return pipes


def write_results(filepath: str,
                  nodes: List[Node],
                  pipes: List[Pipe],
                  result,   # SolverResult
                  gas,      # GasProperties
                  sheet: str = 'RESULTS') -> None:
    """
    Write solver results to a new sheet in the Excel workbook.

    Creates two tables:
        1. Node results  — nodal pressures, pressure margins
        2. Pipe results  — flows, velocities, Re, f, pressure gradient, loading

    Parameters
    ----------
    filepath : str         path to the .xlsx file (modified in-place)
    nodes    : list        Node objects (with solved pressures)
    pipes    : list        Pipe objects (with computed flows)
    result   : SolverResult
    gas      : GasProperties
    sheet    : str         output sheet name
    """
    from postprocess import compute_pipe_results, compute_node_results

    pipe_df = compute_pipe_results(pipes, nodes, result, gas)
    node_df = compute_node_results(nodes, result)

    wb = load_workbook(filepath)

    # Remove old RESULTS sheet if it exists
    if sheet in wb.sheetnames:
        del wb[sheet]

    ws = wb.create_sheet(sheet)
    ws.sheet_properties.tabColor = "375623"

    # ── Colour palette ──
    NAVY      = "1F4E79"
    BLUE      = "2E75B6"
    LIGHT_BLU = "D6E4F0"
    GREEN_DK  = "375623"
    LIGHT_GRN = "E2EFDA"
    ORANGE    = "C55A11"
    LIGHT_ORN = "FCE4D6"
    RED       = "C00000"
    LIGHT_RED = "FCE4D6"
    WHITE     = "FFFFFF"
    LGRAY     = "F2F2F2"
    VLIGHT    = "EBF3FB"

    s = Side(style="thin", color="CCCCCC")
    bdr = Border(top=s, bottom=s, left=s, right=s)

    def hdr_cell(ws, row, col, value, bg=NAVY, fg=WHITE, size=10):
        c = ws.cell(row=row, column=col)
        c.value = value
        c.font = Font(name="Arial", bold=True, size=size, color=fg)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = bdr
        return c

    def data_cell(ws, row, col, value, bg=WHITE, fg="2c2c2a",
                  bold=False, fmt=None, align="center"):
        c = ws.cell(row=row, column=col)
        c.value = value
        c.font = Font(name="Arial", size=10, color=fg, bold=bold)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = bdr
        if fmt:
            c.number_format = fmt
        return c

    # ── Banner ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = (f"PE80 GAS NETWORK — SOLVER RESULTS   |   "
               f"{'CONVERGED' if result.converged else 'NOT CONVERGED'}   |   "
               f"{result.iterations} iterations   |   "
               f"max|F| = {result.final_residual:.3e} Sm³/s")
    c.font = Font(name="Arial", bold=True, size=12, color=WHITE)
    c.fill = PatternFill("solid", fgColor=GREEN_DK if result.converged else RED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # ── NODE RESULTS TABLE ────────────────────────────────────────────────────
    ws.merge_cells("A3:G3")
    c = ws["A3"]
    c.value = "NODE RESULTS"
    c.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 20

    node_headers = [
        "Node_ID", "Label", "Type",
        "Solved_P\n[kPa(g)]", "Min_P\n[kPa(g)]",
        "Margin\n[kPa]", "Status"
    ]
    for j, h in enumerate(node_headers):
        hdr_cell(ws, 4, j+1, h, bg=NAVY)
    ws.row_dimensions[4].height = 32

    for i, row in node_df.iterrows():
        r = i + 5
        bg = VLIGHT if i % 2 == 0 else WHITE
        margin = row.get('Margin_kPa', 0)
        status = row.get('Status', '✓')

        data_cell(ws, r, 1, row['Node_ID'],   bg=bg, bold=True, fg=NAVY, align="left")
        data_cell(ws, r, 2, row['Label'],      bg=bg, align="left")
        data_cell(ws, r, 3, row['Type'],       bg=bg)
        data_cell(ws, r, 4, row.get('P_kPag', 0),   bg=bg, fmt="#,##0.1")
        data_cell(ws, r, 5, row.get('MinP_kPag', 0), bg=bg, fmt="#,##0.1")

        # Margin — red if negative
        m_bg = LIGHT_RED if margin < 0 else bg
        m_fg = RED if margin < 0 else "2c2c2a"
        data_cell(ws, r, 6, margin, bg=m_bg, fg=m_fg, fmt="#,##0.1")

        s_bg = LIGHT_RED if '✗' in str(status) else LIGHT_GRN if '✓' in str(status) else bg
        s_fg = RED if '✗' in str(status) else GREEN_DK if '✓' in str(status) else "2c2c2a"
        data_cell(ws, r, 7, status, bg=s_bg, fg=s_fg, bold=True)
        ws.row_dimensions[r].height = 18

    # ── PIPE RESULTS TABLE ────────────────────────────────────────────────────
    pipe_start_row = 5 + len(node_df) + 3

    ws.merge_cells(f"A{pipe_start_row}:L{pipe_start_row}")
    c = ws.cell(pipe_start_row, 1)
    c.value = "PIPE RESULTS"
    c.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    c.fill = PatternFill("solid", fgColor=ORANGE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[pipe_start_row].height = 20

    pipe_headers = [
        "Pipe_ID", "From", "To",
        "Q\n[Sm³/h]", "Direction",
        "v\n[m/s]", "Re\n[-]",
        "f\n[-]",
        "ΔP\n[mbar]",
        "Grad\n[mbar/m]",
        "Loading\n[%]",
        "Status"
    ]
    hdr_row = pipe_start_row + 1
    for j, h in enumerate(pipe_headers):
        hdr_cell(ws, hdr_row, j+1, h, bg=ORANGE)
    ws.row_dimensions[hdr_row].height = 32

    for i, row in pipe_df.iterrows():
        r = hdr_row + 1 + i
        bg = VLIGHT if i % 2 == 0 else WHITE
        loading = row.get('Loading_pct', 0)
        v       = row.get('v_ms', 0)
        status  = row.get('Status', '✓')

        # Loading colour: green < 70%, amber 70–90%, red > 90%
        if loading > 90:
            l_bg, l_fg = LIGHT_RED, RED
        elif loading > 70:
            l_bg, l_fg = "FFF2CC", "7F6000"
        else:
            l_bg, l_fg = LIGHT_GRN, GREEN_DK

        data_cell(ws, r,  1, row['Pipe_ID'],       bg=bg, bold=True, fg=ORANGE, align="left")
        data_cell(ws, r,  2, row['From'],           bg=bg, align="left")
        data_cell(ws, r,  3, row['To'],             bg=bg, align="left")
        data_cell(ws, r,  4, row.get('Q_sm3h', 0), bg=bg, fmt="#,##0.1")
        data_cell(ws, r,  5, row.get('Direction','→'), bg=bg)
        data_cell(ws, r,  6, v,                    bg=bg, fmt="0.00")
        data_cell(ws, r,  7, row.get('Re', 0),     bg=bg, fmt="#,##0")
        data_cell(ws, r,  8, row.get('f', 0),      bg=bg, fmt="0.00000")
        data_cell(ws, r,  9, row.get('dP_mbar', 0), bg=bg, fmt="#,##0.1")
        data_cell(ws, r, 10, row.get('Grad_mbar_m', 0), bg=bg, fmt="0.000")
        data_cell(ws, r, 11, loading,              bg=l_bg, fg=l_fg, fmt="0.0")
        data_cell(ws, r, 12, status,               bg=bg)
        ws.row_dimensions[r].height = 18

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [10, 16, 16, 12, 10, 10, 12, 10, 12, 12, 10, 14]
    for j, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(j+1)].width = w

    # ── Mass balance summary ──────────────────────────────────────────────────
    mb_row = hdr_row + 1 + len(pipe_df) + 2
    from network import build_incidence_matrix
    import numpy as np
    _A       = build_incidence_matrix(nodes, pipes)
    _src_idx = [i for i, n in enumerate(nodes) if n.is_source()]
    _Q       = np.array([p.flow for p in pipes])
    total_supply = float(np.sum(_A[_src_idx, :] @ _Q)) * 3600
    total_demand = sum(n.demand * 3600 for n in nodes if n.is_offtake())
    balance      = total_supply - total_demand

    ws.merge_cells(f"A{mb_row}:D{mb_row}")
    ws.cell(mb_row, 1).value = "MASS BALANCE"
    ws.cell(mb_row, 1).font  = Font(name="Arial", bold=True, size=10, color=WHITE)
    ws.cell(mb_row, 1).fill  = PatternFill("solid", fgColor=NAVY)
    ws.cell(mb_row, 1).alignment = Alignment(horizontal="center")

    for label, value, col in [
        ("Total supply [Sm³/h]",  total_supply, 5),
        ("Total demand [Sm³/h]",  total_demand, 7),
        ("Balance [Sm³/h]",       balance,      9),
    ]:
        ws.cell(mb_row, col).value = label
        ws.cell(mb_row, col).font  = Font(name="Arial", bold=True, size=10)
        ws.cell(mb_row, col+1).value = round(value, 3)
        ws.cell(mb_row, col+1).font  = Font(name="Arial", size=10,
                                             color=RED if abs(balance) > 0.1 else GREEN_DK)

    wb.save(filepath)
    log.info(f"Results written to sheet '{sheet}' in {filepath}")
